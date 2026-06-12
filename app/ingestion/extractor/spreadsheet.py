"""Excel(xlsx)/CSV 텍스트 추출 (FR-002) — 시트→자연어 직렬화 [Pipeline].

작성자 : 최태성
담당 영역 : ingestion

수치 표를 LLM 이 맥락으로 이해할 수 있도록 행을 ``헤더: 값`` 형태로 직렬화한다(첫 행을
헤더로 가정). 라이브러리는 지연 import 한다.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable


def extract_xlsx_text(content: bytes) -> str:
    """xlsx 바이너리의 모든 시트를 자연어 직렬화 텍스트로 변환한다."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        blocks = [
            _serialize_sheet(sheet.title, sheet.iter_rows(values_only=True))
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()
    return "\n\n".join(block for block in blocks if block).strip()


def extract_csv_text(content: bytes) -> str:
    """CSV 바이너리를 자연어 직렬화 텍스트로 변환한다(UTF-8 BOM 허용)."""
    text = content.decode("utf-8-sig", errors="replace")
    rows = csv.reader(io.StringIO(text))
    return _serialize_sheet("CSV", rows)


def _serialize_sheet(name: str, raw_rows: Iterable[tuple | list]) -> str:
    """시트 행들을 ``Sheet: <name>`` + ``헤더: 값`` 라인으로 직렬화한다.

    빈 행은 제외한다. 첫 비어있지 않은 행을 헤더로 사용하고, 이후 행은 헤더와 매핑한다.
    데이터 행이 없으면(헤더만) 헤더 값을 한 줄로 포함한다.
    """
    rows = [_normalize_row(row) for row in raw_rows]
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        return ""

    header = rows[0]
    lines = [f"Sheet: {name}"]
    for row in rows[1:]:
        cells = []
        for index, value in enumerate(row):
            if not value:
                continue
            key = header[index] if index < len(header) and header[index] else f"col{index + 1}"
            cells.append(f"{key}: {value}")
        if cells:
            lines.append(", ".join(cells))
    if len(rows) == 1:
        lines.append(", ".join(cell for cell in header if cell))
    return "\n".join(lines)


def _normalize_row(row: tuple | list) -> list[str]:
    """행 셀을 문자열로 정규화(None→'' , 공백 trim)."""
    return ["" if cell is None else str(cell).strip() for cell in row]
